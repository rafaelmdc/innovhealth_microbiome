from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook

from database.models import (
    AlphaMetric,
    BetaMetric,
    Comparison,
    Group,
    ImportBatch,
    MetadataValue,
    MetadataVariable,
    Taxon,
    QualitativeFinding,
    QuantitativeFinding,
    Study,
)

from .services import build_preview, run_import


class ImportServiceTests(TestCase):
    def setUp(self):
        self.study = Study.objects.create(title='Study A', doi='10.1000/example')
        self.group_a = Group.objects.create(study=self.study, name='Case')
        self.group_b = Group.objects.create(study=self.study, name='Control')
        self.comparison = Comparison.objects.create(
            study=self.study,
            group_a=self.group_a,
            group_b=self.group_b,
            label='Case vs control',
        )
        self.taxon = Taxon.objects.create(
            ncbi_taxonomy_id=100,
            scientific_name='Taxon A',
            rank='species',
        )
        self.metadata_variable = MetadataVariable.objects.create(
            name='smoking_status',
            display_name='Smoking Status',
            value_type=MetadataVariable.ValueType.TEXT,
        )

    def test_study_preview_reports_duplicate_doi(self):
        preview = build_preview(
            file_name='studies.csv',
            content='doi,title\n10.1000/example,Duplicate Study\n',
            import_type='study',
            batch_name='Study batch',
        )

        self.assertEqual(preview.valid_rows, [])
        self.assertEqual(len(preview.duplicates), 1)

    def test_group_preview_requires_existing_study(self):
        preview = build_preview(
            file_name='groups.csv',
            content='study_doi,study_title,name\n10.9999/missing,,Cohort X\n',
            import_type='group',
            batch_name='Group batch',
        )

        self.assertEqual(preview.valid_rows, [])
        self.assertEqual(len(preview.errors), 1)

    def test_comparison_import_creates_row(self):
        preview = build_preview(
            file_name='comparisons.csv',
            content=(
                'study_doi,study_title,group_a_name,group_b_name,label\n'
                '10.1000/example,,Case,Control,Case vs control import\n'
            ),
            import_type='comparison',
            batch_name='Comparison batch',
        ).to_dict()

        batch = run_import(preview)

        self.assertEqual(batch.status, ImportBatch.Status.COMPLETED)
        self.assertTrue(
            Comparison.objects.filter(
                study=self.study,
                group_a=self.group_a,
                group_b=self.group_b,
                label='Case vs control import',
            ).exists()
        )

    def test_metadata_value_import_creates_value(self):
        preview = build_preview(
            file_name='metadata_values.csv',
            content=(
                'study_doi,study_title,group_name,variable_name,value_text\n'
                '10.1000/example,,Case,smoking_status,never\n'
            ),
            import_type='metadata_value',
            batch_name='Metadata value batch',
        ).to_dict()

        batch = run_import(preview)
        metadata_value = MetadataValue.objects.get(group=self.group_a, variable=self.metadata_variable)

        self.assertEqual(batch.status, ImportBatch.Status.COMPLETED)
        self.assertEqual(metadata_value.value_text, 'never')

    def test_qualitative_finding_import_creates_batch_link(self):
        preview = build_preview(
            file_name='qualitative_findings.csv',
            content=(
                'study_doi,study_title,group_a_name,group_b_name,comparison_label,taxon_scientific_name,direction,source\n'
                '10.1000/example,,Case,Control,Case vs control,Taxon A,enriched,Table 2\n'
            ),
            import_type='qualitative_finding',
            batch_name='Qualitative batch',
        ).to_dict()

        batch = run_import(preview)
        finding = QualitativeFinding.objects.get(comparison=self.comparison, taxon=self.taxon)

        self.assertEqual(batch.status, ImportBatch.Status.COMPLETED)
        self.assertEqual(finding.import_batch, batch)
        self.assertEqual(finding.direction, QualitativeFinding.Direction.ENRICHED)

    def test_quantitative_finding_preview_requires_numeric_value(self):
        preview = build_preview(
            file_name='quantitative_findings.csv',
            content=(
                'study_doi,study_title,group_name,taxon_scientific_name,value_type,value,source\n'
                '10.1000/example,,Case,Taxon A,relative_abundance,not-a-number,Table 3\n'
            ),
            import_type='quantitative_finding',
            batch_name='Quantitative batch',
        )

        self.assertEqual(preview.valid_rows, [])
        self.assertEqual(len(preview.errors), 1)

    def test_quantitative_finding_import_creates_batch_link(self):
        preview = build_preview(
            file_name='quantitative_findings.csv',
            content=(
                'study_doi,study_title,group_name,taxon_scientific_name,value_type,value,source\n'
                '10.1000/example,,Case,Taxon A,relative_abundance,0.42,Table 3\n'
            ),
            import_type='quantitative_finding',
            batch_name='Quantitative batch',
        ).to_dict()

        batch = run_import(preview)
        finding = QuantitativeFinding.objects.get(group=self.group_a, taxon=self.taxon)

        self.assertEqual(batch.status, ImportBatch.Status.COMPLETED)
        self.assertEqual(finding.import_batch, batch)
        self.assertEqual(finding.value, 0.42)

    def test_organism_import_creates_import_batch_and_records(self):
        preview = build_preview(
            file_name='taxa.csv',
            content=(
                'ncbi_taxonomy_id,scientific_name,rank,notes\n'
                '101,Faecalibacterium prausnitzii,species,Important commensal\n'
            ),
            import_type='taxon',
            batch_name='Taxon batch',
        )
        preview_row = preview.valid_rows[0]

        batch = run_import(preview.to_dict())

        self.assertEqual(ImportBatch.objects.count(), 1)
        self.assertEqual(batch.status, ImportBatch.Status.COMPLETED)
        self.assertEqual(batch.success_count, 1)
        self.assertEqual(batch.error_count, 0)
        self.assertTrue(
            Taxon.objects.filter(
                ncbi_taxonomy_id=preview_row['ncbi_taxonomy_id'],
                scientific_name=preview_row['scientific_name'],
            ).exists()
        )

    def test_taxon_preview_includes_resolution_metadata_and_lineage(self):
        preview = build_preview(
            file_name='taxa.csv',
            content=(
                'scientific_name,rank,notes\n'
                'Faecalibacterium prausnitzii,species,Important commensal\n'
            ),
            import_type='taxon',
            batch_name='Taxon preview batch',
        )

        self.assertEqual(preview.errors, [])
        self.assertEqual(len(preview.valid_rows), 1)
        row = preview.valid_rows[0]
        self.assertEqual(row['scientific_name'], 'Faecalibacterium prausnitzii')
        self.assertFalse(row['review_required'])
        self.assertTrue(row['lineage'])
        self.assertIn('Faecalibacterium prausnitzii', row['lineage_summary'])
        self.assertTrue(row['resolution_message'])

    def test_taxon_preview_page_shows_resolver_states(self):
        user_model = get_user_model()
        staff_user = user_model.objects.create_user(
            username='importstaff',
            password='testpass123',
            is_staff=True,
        )
        self.client.login(username='importstaff', password='testpass123')
        session = self.client.session
        session['imports_preview'] = {
            'batch_name': 'Taxon UI batch',
            'import_type': 'taxon',
            'required_columns': ['scientific_name', 'rank'],
            'file_name': 'taxa.csv',
            'total_rows': 3,
            'valid_rows': [
                {
                    'row_number': 2,
                    'scientific_name': 'Auto Taxon',
                    'rank': 'species',
                    'ncbi_taxonomy_id': 1,
                    'resolution_status': 'resolved_exact_scientific',
                    'resolution_message': 'resolved exact scientific',
                    'review_required': False,
                    'resolver_source': 'taxonbridge_name',
                    'lineage_summary': 'root > Auto Taxon',
                },
                {
                    'row_number': 3,
                    'scientific_name': 'Review Taxon',
                    'rank': 'species',
                    'ncbi_taxonomy_id': '',
                    'resolution_status': 'manual_review_required',
                    'resolution_message': 'manual review required',
                    'review_required': True,
                    'resolver_source': 'taxonbridge_name',
                    'lineage_summary': 'Review Taxon',
                },
                {
                    'row_number': 4,
                    'scientific_name': 'Fallback Taxon',
                    'rank': 'genus',
                    'ncbi_taxonomy_id': '',
                    'resolution_status': 'taxonbridge_unavailable',
                    'resolution_message': 'Taxonomy DB not found',
                    'review_required': False,
                    'resolver_source': 'local_fallback',
                    'lineage_summary': 'Fallback Taxon',
                },
            ],
            'errors': [],
            'duplicates': [],
        }
        session.save()

        response = self.client.get(reverse('imports:preview'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Auto-resolved')
        self.assertContains(response, 'Review required')
        self.assertContains(response, 'Fallback local')
        self.assertContains(response, 'resolver-chip-auto')
        self.assertContains(response, 'resolver-chip-review')
        self.assertContains(response, 'resolver-chip-fallback')
        self.assertContains(response, 'Taxonomy DB not found')

    def test_alpha_metric_import_sets_import_batch(self):
        preview = build_preview(
            file_name='alpha_metrics.csv',
            content=(
                'study_doi,study_title,group_name,metric,value,source\n'
                '10.1000/example,,Case,shannon,3.82,Table 4\n'
            ),
            import_type='alpha_metric',
            batch_name='Alpha metric batch',
        ).to_dict()

        batch = run_import(preview)
        metric = AlphaMetric.objects.get(group=self.group_a, metric='shannon')

        self.assertEqual(batch.status, ImportBatch.Status.COMPLETED)
        self.assertEqual(metric.import_batch, batch)
        self.assertEqual(metric.value, 3.82)

    def test_beta_metric_import_sets_import_batch(self):
        preview = build_preview(
            file_name='beta_metrics.csv',
            content=(
                'study_doi,study_title,group_a_name,group_b_name,comparison_label,metric,value,source\n'
                '10.1000/example,,Case,Control,Case vs control,bray_curtis,0.37,Figure 2\n'
            ),
            import_type='beta_metric',
            batch_name='Beta metric batch',
        ).to_dict()

        batch = run_import(preview)
        metric = BetaMetric.objects.get(comparison=self.comparison, metric='bray_curtis')

        self.assertEqual(batch.status, ImportBatch.Status.COMPLETED)
        self.assertEqual(metric.import_batch, batch)
        self.assertEqual(metric.value, 0.37)

    def _build_workbook_bytes(self):
        workbook = Workbook()
        paper_sheet = workbook.active
        paper_sheet.title = 'Paper'
        paper_sheet.append(
            ['paper_id', 'doi', 'authors', 'year', 'title', 'country', 'topic', 'status', 'reviwer', 'notes']
        )
        paper_sheet.append(
            ['paper-1', '10.2000/workbook', 'A. Author, B. Author', 2024, 'Workbook Study', 'Portugal', 'IBD', 'complete', 'Rafael', 'Paper note']
        )
        paper_sheet.append(
            ['paper-2', '10.2000/skip', 'C. Author', 2023, 'Skipped Study', 'Spain', 'Control', 'hold', 'Rafael', 'Skip me']
        )

        group_sheet = workbook.create_sheet('groups')
        group_sheet.append(
            ['group_id', 'paper_id', 'group_name_as_written', 'condition', 'group_type', 'body_site', 'sample_size', 'age', 'women_percent', 'age2', 'where_found', 'notes']
        )
        group_sheet.append(['g1', 'paper-1', 'Cases', 'IBD', 'case', 'gut', 12, 40.5, 60, 'adult', 'Methods', 'Case group'])
        group_sheet.append(['g2', 'paper-1', 'Controls', 'Healthy', 'control', 'gut', 10, 38, 55, '', 'Methods', 'Control group'])
        group_sheet.append(['g3', 'paper-2', 'Skipped Group', 'Other', 'other', 'gut', 5, '', '', '', '', 'Should skip'])

        comparison_sheet = workbook.create_sheet('comparisons')
        comparison_sheet.append(
            ['comparison_id', 'paper_id', 'target_group_id', 'reference_group_id', 'target_condition', 'reference_condition', 'comparison_type', 'notes']
        )
        comparison_sheet.append(['c1', 'paper-1', 'g1', 'g2', 'IBD', 'Healthy', 'case_vs_control', 'Comparison note'])

        qualitative_sheet = workbook.create_sheet('qualitative_findings')
        qualitative_sheet.append(
            ['finding_id', 'paper_id', 'comparison_id', 'organism_id', 'organism_as_writiten', 'direction', 'finding_type', 'where_found', 'notes']
        )
        qualitative_sheet.append(['f1', 'paper-1', 'c1', 'o1', 'Blautia sp.', 'increased_in_target', 'relative_direction', 'Table 2', 'Finding note'])

        quantitative_sheet = workbook.create_sheet('quantitive_findings')
        quantitative_sheet.append(
            ['quant_finding_id', 'paper_id', 'group_id', 'organism_id', 'value_type', 'unit', 'value', 'where_found', 'notes']
        )
        quantitative_sheet.append(['q1', 'paper-1', 'g1', 'o1', 'relative_abundance', '%', 0.42, 'Table 3', 'Quant note'])

        diversity_sheet = workbook.create_sheet('diversity_metrics')
        diversity_sheet.append(
            ['diversity_id', 'paper_id', 'comparison_id', 'group_id', 'diversity_category', 'metric_as_written', 'value', 'unit', 'where_found', 'notes']
        )
        diversity_sheet.append(['d1', 'paper-1', '', 'g1', 'alpha', 'shannon', 3.4, '', 'Table 4', 'Alpha note'])
        diversity_sheet.append(['d2', 'paper-1', 'c1', '', 'beta', 'bray_curtis', 0.28, '', 'Figure 2', 'Beta note'])

        organism_sheet = workbook.create_sheet('organisms')
        organism_sheet.append(
            ['organism_id', 'organism_as_written', 'suggested_clean_name', 'rank_if_known', 'notes', 'ncbi_id', 'resolved']
        )
        organism_sheet.append(['o1', 'Blautia sp.', 'Blautia', 'genus', 'Taxon note', 1234, 'true'])

        metadata_sheet = workbook.create_sheet('extra_metadata')
        metadata_sheet.append(['paper_id', 'group_id', 'field_name', 'value_as_written', 'unit', 'where_found', 'notes'])
        metadata_sheet.append(['paper-1', 'g1', 'bmi', '24.5', 'kg/m2', 'Table 1', 'BMI note'])

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def test_excel_workbook_preview_and_import_creates_related_records(self):
        preview = build_preview(
            file_name='curation.xlsx',
            content=self._build_workbook_bytes(),
            import_type='excel_workbook',
            batch_name='Workbook batch',
        )

        self.assertEqual(preview.import_type, 'excel_workbook')
        self.assertEqual(len(preview.errors), 0)
        self.assertEqual(len(preview.sections), 10)
        self.assertEqual(len(preview.skipped_rows), 2)

        batch = run_import(preview.to_dict())

        self.assertEqual(batch.status, ImportBatch.Status.COMPLETED)
        study = Study.objects.get(doi='10.2000/workbook')
        cases = Group.objects.get(study=study, name='Cases')
        controls = Group.objects.get(study=study, name='Controls')
        comparison = Comparison.objects.get(
            study=study,
            group_a=cases,
            group_b=controls,
            label='Cases vs Controls (case_vs_control)',
        )
        taxon = Taxon.objects.get(scientific_name='Blautia')
        qualitative = QualitativeFinding.objects.get(comparison=comparison, taxon=taxon)
        quantitative = QuantitativeFinding.objects.get(group=cases, taxon=taxon)
        alpha = AlphaMetric.objects.get(group=cases, metric='shannon')
        beta = BetaMetric.objects.get(comparison=comparison, metric='bray_curtis')
        bmi_variable = MetadataVariable.objects.get(name='bmi')
        bmi_value = MetadataValue.objects.get(group=cases, variable=bmi_variable)
        age_variable = MetadataVariable.objects.get(name='age')
        age_value = MetadataValue.objects.get(group=cases, variable=age_variable)
        women_variable = MetadataVariable.objects.get(name='women_percent')
        women_value = MetadataValue.objects.get(group=cases, variable=women_variable)

        self.assertIn('Authors: A. Author, B. Author', study.notes)
        self.assertEqual(qualitative.direction, QualitativeFinding.Direction.ENRICHED)
        self.assertEqual(qualitative.import_batch, batch)
        self.assertEqual(quantitative.value, 0.42)
        self.assertEqual(alpha.import_batch, batch)
        self.assertEqual(beta.import_batch, batch)
        self.assertEqual(bmi_variable.value_type, MetadataVariable.ValueType.TEXT)
        self.assertEqual(bmi_value.value_text, '24.5')
        self.assertEqual(age_variable.value_type, MetadataVariable.ValueType.TEXT)
        self.assertEqual(age_value.value_text, '40.5')
        self.assertEqual(women_variable.value_type, MetadataVariable.ValueType.TEXT)
        self.assertEqual(women_value.value_text, '60')
        self.assertFalse(Study.objects.filter(doi='10.2000/skip').exists())

    def test_excel_workbook_aliases_skip_unresolved_and_upsert_records(self):
        study = Study.objects.create(
            doi='10.4000/upsert',
            title='Workbook Alias Study',
            country='Old Country',
            notes='Old study note',
        )
        cases = Group.objects.create(study=study, name='Cases', notes='Old case note')
        controls = Group.objects.create(study=study, name='Controls', notes='Old control note')
        comparison = Comparison.objects.create(
            study=study,
            group_a=cases,
            group_b=controls,
            label='Cases vs Controls (case_vs_control)',
            notes='Old comparison note',
        )
        taxon = Taxon.objects.create(
            ncbi_taxonomy_id=2222,
            scientific_name='Resolved Taxon',
            rank='old_rank',
            notes='Old taxon note',
        )
        QuantitativeFinding.objects.create(
            group=cases,
            taxon=taxon,
            value_type=QuantitativeFinding.ValueType.RELATIVE_ABUNDANCE,
            value=0.1,
            unit='%',
            source='Table 3',
            notes='Old quantitative note',
        )

        workbook = Workbook()
        paper_sheet = workbook.active
        paper_sheet.title = 'Paper'
        paper_sheet.append(['paper_id', 'doi', 'authors', 'year', 'title', 'country', 'topic', 'status', 'reviwer', 'notes'])
        paper_sheet.append(['paper-1', '10.4000/upsert', 'A. Author', 2024, 'Workbook Alias Study', 'Portugal', 'IBD', 'complete', 'Rafael', 'Updated study note'])
        paper_sheet.append(['paper-2', '10.4000/skip', 'B. Author', 2024, 'Skipped Alias Study', 'Spain', 'IBD', 'to review', 'Rafael', 'Should skip'])

        group_sheet = workbook.create_sheet('groups')
        group_sheet.append(['group_id', 'paper_id', 'group_name_as_written', 'condition', 'group_type', 'body_site', 'sample_size', 'age', 'women_percent', 'age2', 'where_found', 'notes'])
        group_sheet.append(['g1', 'paper-1', 'Cases', 'IBD', 'case', 'gut', 12, 40.5, 60, '', 'Methods', 'Updated case note'])
        group_sheet.append(['g2', 'paper-1', 'Controls', 'Healthy', 'control', 'gut', 10, 55, 55, '', 'Methods', 'Updated control note'])
        group_sheet.append(['g3', 'paper-2', 'Skipped Group', 'Other', 'other', 'gut', 5, '', '', '', '', 'Should skip'])

        comparison_sheet = workbook.create_sheet('comparisons')
        comparison_sheet.append(['comparison_id', 'paper_id', 'target_group_id', 'reference_group_id', 'comparison_type', 'notes'])
        comparison_sheet.append(['c1', 'paper-1', 'g1', 'g2', 'case_vs_control', 'Updated comparison note'])

        qualitative_sheet = workbook.create_sheet('qualitative_findings')
        qualitative_sheet.append(['finding_id', 'paper_id', 'comparison_id', 'organism_id', 'organism_as_writiten', 'direction', 'finding_type', 'where_found', 'notes'])
        qualitative_sheet.append(['f1', 'paper-1', 'c1', 'o1', 'Resolved Taxon', 'increased_in_target', 'relative_direction', 'Table 2', 'Updated qualitative note'])
        qualitative_sheet.append(['f2', 'paper-1', 'c1', 'o2', 'Unresolved Taxon', 'decreased_in_target', 'relative_direction', 'Table 5', 'Should skip'])

        quantitative_sheet = workbook.create_sheet('quantitive_findings')
        quantitative_sheet.append(['quant_finding_id', 'paper_id', 'group_id', 'organism_id', 'value_type', 'unit', 'value', 'where_found', 'notes'])
        quantitative_sheet.append(['q1', 'paper-1', 'g1', 'o1', 'relative_abundance', '%', 0.42, 'Table 3', 'Updated quantitative note'])
        quantitative_sheet.append(['q2', 'paper-1', 'g1', 'o2', 'relative_abundance', '%', 0.33, 'Table 4', 'Should skip'])

        organism_sheet = workbook.create_sheet('organisms')
        organism_sheet.append(['organism_id', 'organism_as_written', 'suggested_clean_name', 'rank_if_known', 'notes', 'ncbi_id', 'resolved'])
        organism_sheet.append(['1', '', '', '', '', '', ''])
        organism_sheet.append(['o1', 'Resolved Taxon', 'Resolved Taxon', 'species', 'Updated taxon note', 2222, 'true'])
        organism_sheet.append(['o2', 'Unresolved Taxon', '', '', 'No match', '', 'false'])

        output = BytesIO()
        workbook.save(output)

        preview = build_preview(
            file_name='alias_workbook.xlsx',
            content=output.getvalue(),
            import_type='excel_workbook',
            batch_name='Alias workbook batch',
        )

        self.assertEqual(preview.import_type, 'excel_workbook')
        self.assertEqual(preview.errors, [])
        self.assertEqual(preview.duplicates, [])
        self.assertTrue(any('paper.status is "to review"' in row['message'] for row in preview.skipped_rows))
        self.assertTrue(any('taxon o2 is not resolved' in row['message'] for row in preview.skipped_rows))

        batch = run_import(preview.to_dict())

        study.refresh_from_db()
        cases.refresh_from_db()
        comparison.refresh_from_db()
        taxon.refresh_from_db()
        quantitative = QuantitativeFinding.objects.get(group=cases, taxon=taxon, source='Table 3')
        qualitative = QualitativeFinding.objects.get(comparison=comparison, taxon=taxon, source='Table 2')

        self.assertEqual(batch.status, ImportBatch.Status.COMPLETED)
        self.assertEqual(study.country, 'Portugal')
        self.assertIn('Updated study note', study.notes)
        self.assertEqual(cases.notes, 'Updated case note\nWhere found: Methods')
        self.assertIn('Updated comparison note', comparison.notes)
        self.assertEqual(taxon.rank, 'species')
        self.assertIn('Updated taxon note', taxon.notes)
        self.assertEqual(quantitative.value, 0.42)
        self.assertEqual(quantitative.import_batch, batch)
        self.assertEqual(qualitative.import_batch, batch)
        self.assertFalse(Taxon.objects.filter(scientific_name='Unresolved Taxon').exists())

    def test_excel_workbook_missing_age_and_women_percent_are_imported_as_na(self):
        workbook = Workbook()
        paper_sheet = workbook.active
        paper_sheet.title = 'Paper'
        paper_sheet.append(['paper_id', 'doi', 'authors', 'year', 'title', 'country', 'topic', 'status', 'reviwer', 'notes'])
        paper_sheet.append(['paper-1', '10.5000/na', 'A. Author', 2024, 'NA Workbook Study', 'Portugal', 'IBD', 'complete', 'Rafael', ''])

        group_sheet = workbook.create_sheet('groups')
        group_sheet.append(['group_id', 'paper_id', 'group_name_as_written', 'condition', 'group_type', 'body_site', 'sample_size', 'age', 'women_percent', 'age2', 'where_found', 'notes'])
        group_sheet.append(['g1', 'paper-1', 'Cases', 'IBD', 'case', 'gut', 12, '', '', '', 'Methods', ''])

        output = BytesIO()
        workbook.save(output)

        preview = build_preview(
            file_name='na_workbook.xlsx',
            content=output.getvalue(),
            import_type='excel_workbook',
            batch_name='NA workbook batch',
        )

        self.assertEqual(preview.errors, [])

        batch = run_import(preview.to_dict())
        study = Study.objects.get(doi='10.5000/na')
        group = Group.objects.get(study=study, name='Cases')
        age_variable = MetadataVariable.objects.get(name='age')
        women_variable = MetadataVariable.objects.get(name='women_percent')
        age_value = MetadataValue.objects.get(group=group, variable=age_variable)
        women_value = MetadataValue.objects.get(group=group, variable=women_variable)

        self.assertEqual(batch.status, ImportBatch.Status.COMPLETED)
        self.assertEqual(age_variable.value_type, MetadataVariable.ValueType.TEXT)
        self.assertEqual(women_variable.value_type, MetadataVariable.ValueType.TEXT)
        self.assertEqual(age_value.value_text, 'NA')
        self.assertEqual(women_value.value_text, 'NA')

    def test_excel_workbook_long_where_found_is_moved_to_notes(self):
        workbook = Workbook()
        paper_sheet = workbook.active
        paper_sheet.title = 'Paper'
        paper_sheet.append(['paper_id', 'doi', 'authors', 'year', 'title', 'country', 'topic', 'status', 'reviwer', 'notes'])
        paper_sheet.append(['paper-1', '10.6000/source', 'A. Author', 2024, 'Long Source Study', 'Portugal', 'IBD', 'complete', 'Rafael', ''])

        group_sheet = workbook.create_sheet('groups')
        group_sheet.append(['group_id', 'paper_id', 'group_name_as_written', 'condition', 'group_type', 'body_site', 'sample_size', 'age', 'women_percent', 'age2', 'where_found', 'notes'])
        group_sheet.append(['g1', 'paper-1', 'Cases', 'IBD', 'case', 'gut', 12, '', '', '', 'Methods', ''])
        group_sheet.append(['g2', 'paper-1', 'Controls', 'Healthy', 'control', 'gut', 10, '', '', '', 'Methods', ''])

        comparison_sheet = workbook.create_sheet('comparisons')
        comparison_sheet.append(['comparison_id', 'paper_id', 'target_group_id', 'reference_group_id', 'comparison_type', 'notes'])
        comparison_sheet.append(['c1', 'paper-1', 'g1', 'g2', 'case_vs_control', ''])

        organism_sheet = workbook.create_sheet('organisms')
        organism_sheet.append(['organism_id', 'organism_as_written', 'suggested_clean_name', 'rank_if_known', 'notes', 'ncbi_id', 'resolved'])
        organism_sheet.append(['o1', 'Bacteroidaceae', 'Bacteroidaceae', 'family', '', 815, 'true'])

        qualitative_sheet = workbook.create_sheet('qualitative_findings')
        qualitative_sheet.append(['finding_id', 'paper_id', 'comparison_id', 'organism_id', 'organism_as_writiten', 'direction', 'finding_type', 'where_found', 'notes'])
        long_source = 'X' * 300
        qualitative_sheet.append(['f1', 'paper-1', 'c1', 'o1', 'Bacteroidaceae', 'increased_in_target', 'relative_direction', long_source, 'Finding note'])

        output = BytesIO()
        workbook.save(output)

        preview = build_preview(
            file_name='long_source.xlsx',
            content=output.getvalue(),
            import_type='excel_workbook',
            batch_name='Long source batch',
        )

        self.assertEqual(preview.errors, [])

        batch = run_import(preview.to_dict())
        study = Study.objects.get(doi='10.6000/source')
        cases = Group.objects.get(study=study, name='Cases')
        controls = Group.objects.get(study=study, name='Controls')
        comparison = Comparison.objects.get(
            study=study,
            group_a=cases,
            group_b=controls,
            label='Cases vs Controls (case_vs_control)',
        )
        taxon = Taxon.objects.get(scientific_name='Bacteroidaceae')
        finding = QualitativeFinding.objects.get(comparison=comparison, taxon=taxon)

        self.assertEqual(batch.status, ImportBatch.Status.COMPLETED)
        self.assertEqual(len(finding.source), 255)
        self.assertIn('Where found:', finding.notes)
        self.assertIn(long_source, finding.notes)

    def test_excel_workbook_preview_reports_missing_group_reference(self):
        workbook = Workbook()
        paper_sheet = workbook.active
        paper_sheet.title = 'paper'
        paper_sheet.append(['paper_id', 'doi', 'authors', 'year', 'title', 'country', 'topic', 'status', 'reviwer', 'notes'])
        paper_sheet.append(['paper-1', '10.3000/error', '', 2024, 'Broken Workbook', '', '', 'complete', '', ''])

        comparison_sheet = workbook.create_sheet('comparissons')
        comparison_sheet.append(
            ['comparison_id', 'paper_id', 'target_group_id', 'reference_group_id', 'target_condition', 'reference_condition', 'comparison_type', 'notes']
        )
        comparison_sheet.append(['c1', 'paper-1', 'missing-a', 'missing-b', '', '', 'case_vs_control', ''])

        output = BytesIO()
        workbook.save(output)

        preview = build_preview(
            file_name='broken.xlsx',
            content=output.getvalue(),
            import_type='excel_workbook',
            batch_name='Broken workbook batch',
        )

        self.assertEqual(preview.import_type, 'excel_workbook')
        self.assertTrue(any(error['section'] == 'comparison' for error in preview.errors))
        self.assertFalse(any(section['valid_rows'] for section in preview.sections if section['import_type'] == 'comparison'))
