"""Shared parsing, normalization, workbook loading, and lookup helpers."""

from io import BytesIO

from openpyxl import load_workbook

from database.models import Comparison, Group, Study

from .taxonomy import resolve_taxon

from .constants import BOOLEAN_FALSE_VALUES, BOOLEAN_TRUE_VALUES, WORKBOOK_SHEET_ALIASES, WORKBOOK_SHEET_ORDER


def cleaned_row(raw_row):
    """Return a row dict with all values stripped and blanks normalized to empty strings."""
    return {key: (value or '').strip() for key, value in raw_row.items()}


def parse_int(value, field_name):
    """Parse a required integer field and return a `(value, error)` tuple."""
    try:
        return int(value), None
    except (TypeError, ValueError):
        return None, f'{field_name} must be an integer.'


def parse_float(value, field_name):
    """Parse a required float field and return a `(value, error)` tuple."""
    try:
        return float(value), None
    except (TypeError, ValueError):
        return None, f'{field_name} must be a float.'


def parse_optional_int(value, field_name):
    """Parse an optional integer field, treating an empty string as `None`."""
    if value == '':
        return None, None
    return parse_int(value, field_name)


def parse_optional_float(value, field_name):
    """Parse an optional float field, treating an empty string as `None`."""
    if value == '':
        return None, None
    return parse_float(value, field_name)


def parse_optional_bool(value, field_name):
    """Parse an optional boolean field using the importer truthy/falsey vocabulary."""
    if value == '':
        return None, None
    normalized = value.lower()
    if normalized in BOOLEAN_TRUE_VALUES:
        return True, None
    if normalized in BOOLEAN_FALSE_VALUES:
        return False, None
    return None, f'{field_name} must be a boolean.'


def resolve_study(study_doi, study_title):
    """Resolve a study by DOI first, then by title."""
    if study_doi:
        return Study.objects.filter(doi=study_doi).first()
    if study_title:
        return Study.objects.filter(title=study_title).first()
    return None


def resolve_group(study_doi, study_title, group_name):
    """Resolve a group inside a study using the current import reference rules."""
    if not group_name:
        return None
    study = resolve_study(study_doi, study_title)
    if not study:
        return None
    return Group.objects.filter(study=study, name=group_name).select_related('study').first()


def resolve_comparison(study_doi, study_title, group_a_name, group_b_name, label):
    """Resolve a comparison by study, group pair, and label."""
    study = resolve_study(study_doi, study_title)
    if not study:
        return None
    queryset = Comparison.objects.filter(
        study=study,
        group_a__name=group_a_name,
        group_b__name=group_b_name,
    )
    if label:
        comparison = queryset.filter(label=label).select_related('study', 'group_a', 'group_b').first()
        if comparison:
            return comparison
    return queryset.select_related('study', 'group_a', 'group_b').first()


def resolve_taxon_reference(scientific_name, ncbi_taxonomy_id):
    """Resolve a taxon by taxonomy ID first, then by known names."""
    return resolve_taxon(scientific_name, ncbi_taxonomy_id)


def row_requires_study_reference(row, errors, row_number):
    """Ensure a CSV row contains at least one study reference field."""
    if row.get('study_doi') or row.get('study_title'):
        return True
    errors.append(
        {
            'row_number': row_number,
            'message': 'At least one of study_doi or study_title is required.',
        }
    )
    return False


def normalize_workbook_cell(value):
    """Normalize workbook cell values into importer-friendly strings."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'true' if value else 'false'
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, 'g')
    return str(value).strip()


def normalize_workbook_sheet_name(sheet_name):
    """Map workbook sheet names to the importer's canonical sheet keys."""
    return WORKBOOK_SHEET_ALIASES.get(sheet_name.strip().lower())


def load_workbook_rows(content):
    """Load supported workbook sheets into a normalized `{fieldnames, rows}` structure."""
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f'Could not read Excel workbook: {exc}') from exc

    sheets = {}
    for sheet_name in workbook.sheetnames:
        canonical_sheet_name = normalize_workbook_sheet_name(sheet_name)
        if canonical_sheet_name not in WORKBOOK_SHEET_ORDER:
            continue

        worksheet = workbook[sheet_name]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration:
            sheets[canonical_sheet_name] = {'fieldnames': [], 'rows': []}
            continue

        fieldnames = [normalize_workbook_cell(value) for value in header_row]
        rows = []
        for row_number, values in enumerate(iterator, start=2):
            normalized_values = [normalize_workbook_cell(value) for value in values]
            if not any(normalized_values):
                continue
            rows.append(
                {
                    'row_number': row_number,
                    'data': {
                        key: normalized_values[index]
                        for index, key in enumerate(fieldnames)
                        if key
                    },
                }
            )
        sheets[canonical_sheet_name] = {'fieldnames': fieldnames, 'rows': rows}

    return sheets


def combine_note_parts(*parts):
    """Join non-empty note fragments with line breaks."""
    return '\n'.join(part for part in parts if part)


def labeled_note(label, value):
    """Format a note fragment only when the underlying value is present."""
    if not value:
        return ''
    return f'{label}: {value}'


def split_source_and_notes(source, *note_parts, max_length=255):
    """Fit a source string into the model field and move overflow into notes."""
    normalized_source = (source or '').strip()
    notes = [part for part in note_parts if part]
    if len(normalized_source) > max_length:
        notes.append(f'Where found: {normalized_source}')
        normalized_source = normalized_source[:max_length]
    return normalized_source, combine_note_parts(*notes)
